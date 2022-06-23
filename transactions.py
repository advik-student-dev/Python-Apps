"""
-------
Imports
-------
"""

# Imports openpyxl library and required classes
import openpyxl as xl
from openpyxl.chart import BarChart, Reference


# Function to process the workbook / sheet
def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    """
    # Learning about accessing the values in a cell 
    cell = sheet['a1'] - same as below
    cell = sheet.cell(1, 1)
    """

    # Goes through the rows with data in the spreadsheet
    for row in range(2, sheet.max_row + 1):

        # References the cell in each row in the specified range
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9

        # References the cell in which the corrected / calculated
        # will go in
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # Assigning the space (specific cells) in which the data will be placed in
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    # Creates new chart object
    chart = BarChart()

    # Adds the values to the Bar Chart
    chart.add_data(values)

    # Adds the chart to the sheet at e2 cell
    sheet.add_chart(chart, 'e2')

    # Saves the changes
    wb.save(filename)


